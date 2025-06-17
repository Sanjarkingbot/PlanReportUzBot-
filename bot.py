from telegram import Update, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, filters, ConversationHandler
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import asyncio
import re

# --- Переменные окружения ---
BOT_TOKEN = os.getenv("BOT_TOKEN")
CHAT_ID = int(os.getenv("CHAT_ID"))

# --- Клавиатура ---
keyboard = [
    [KeyboardButton("🗓 План на сегодня"), KeyboardButton("📝 Отправить отчёт")],
    [KeyboardButton("🗕 Отчёты по дате"), KeyboardButton("🔍 Поиск по сотруднику")],
    [KeyboardButton("📂 Шаблон отчёта")]
]
reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

TYPING_PLAN = 1
user_states = {}

# --- Excel файл ---
def get_excel():
    file_name = "reports.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Имя", "Username", "Тип", "№", "Задача", "Комментарий"])
        wb.save(file_name)
    return file_name

# --- Команды ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Добро пожаловать! Используйте кнопки или команды.", reply_markup=reply_markup)

async def template(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📂 Шаблон отчёта:\n\n"
        "1. Задача\n"
        "— Комментарий\n\n"
        "2. Задача\n"
        "— Комментарий"
    )

async def plan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_states[user_id] = "ПЛАН"
    await update.message.reply_text(
        "✏️ Введите свой ПЛАН на сегодня в следующем формате:\n\n"
        "📄 Шаблон плана на 2025-06-17:\n\n"
        "1. Задача\n"
        "— Что нужно сделать\n\n"
        "2. Задача\n"
        "— Что планируете достичь\n\n"
        "🔁 Пример:\n"
        "1. Встреча с подрядчиком\n"
        "— Обсудить этапы установки оборудования\n\n"
        "2. Подготовка документов\n"
        "— Завершить пакет для поставщика"
    )
    return TYPING_PLAN

async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_states[user_id] = "ОТЧЁТ"
    await update.message.reply_text(
        "✏️ Введите свой ОТЧЁТ в виде текста, как в шаблоне. Если допустили ошибку — просто отправьте новый отчёт с пометкой ‘исправление’."
    )
    return TYPING_PLAN

async def save_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    user_id = user.id
    text = update.message.text.strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    entry_type = user_states.get(user_id)

    file_name = get_excel()
    wb = load_workbook(file_name)
    ws = wb.active

    entries = re.findall(r"(\d+)\.\s*(.+?)\s*(?:—|-|:)\s*(.+)", text)
    if entries:
        for num, task, comment in entries:
            if comment.strip():
                ws.append([now, user.full_name, user.username, entry_type, num.strip(), task.strip(), comment.strip()])
            else:
                return await update.message.reply_text(f"⚠️ В строке №{num} отсутствует комментарий.")
        wb.save(file_name)
        user_states.pop(user_id, None)
        return await update.message.reply_text("✅ План сохранён. Спасибо!")
    else:
        await update.message.reply_text("⚠️ Формат плана некорректный. Используйте шаблон: /template")
        return TYPING_PLAN

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_states.pop(user_id, None)
    await update.message.reply_text("❌ Операция отменена.")
    return ConversationHandler.END

# --- Напоминания ---
async def send_daily_plan(context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=CHAT_ID, text=
        "📌 План на сегодня:\n"
        "1. Утреннее собрание\n"
        "2. Работы на площадке\n"
        "3. Согласование с поставщиками\n"
        "4. Документооборот\n"
        "5. Отправка отчёта вечером"
    )

async def send_report_reminder(context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(chat_id=CHAT_ID, text="⏰ Напоминание: не забудьте отправить отчёт до 18:00!")

# --- Старт приложения ---
async def on_startup(app):
    print("✅ Бот запущен и готов к работе")

async def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Обработчики
    conv_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^🗓 План на сегодня$"), plan),
            MessageHandler(filters.Regex("^📝 Отправить отчёт$"), report)
        ],
        states={TYPING_PLAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_entry)]},
        fallbacks=[CommandHandler("cancel", cancel)]
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("template", template))
    app.add_handler(conv_handler)

    # Планировщик задач
    scheduler = AsyncIOScheduler()
    scheduler.add_job(send_daily_plan, "cron", hour=7, minute=50, args=[app])
    scheduler.add_job(send_report_reminder, "cron", hour=17, minute=0, args=[app])
    scheduler.start()

    # Webhook
    await app.initialize()
    await app.bot.set_webhook("https://planreportuzbot.onrender.com/webhook")
    await app.start()
    await app.updater.start_webhook(
        listen="0.0.0.0",
        port=int(os.getenv("PORT", "10000")),
        url_path="webhook",
        webhook_url="https://planreportuzbot.onrender.com/webhook",
        on_startup=[on_startup]
    )

if __name__ == "__main__":
    asyncio.run(main())
